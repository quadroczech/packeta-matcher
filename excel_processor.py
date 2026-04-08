import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
from zasilkovna_api import get_packet_data, get_client

MAX_WORKERS = 10  # počet paralelních vláken pro API volání

ORDER_COLUMN_HEADER = "Číslo objednávky"
PRICE_COLUMN_HEADER = "Celková cena CHF"
TRACKING_PATTERN = re.compile(r'^[Zz]?\d{6,15}$')

VAT_RATE = 0.081             # DPH 8,1 %
SHIPPING_IN_CUSTOMS = 9.0    # Zásilkovna vždy deklaruje 9 CHF dopravy v Customs Value (bez DPH)
SHIPPING_REAL_GROSS = 9.99   # Reálná cena dopravy včetně DPH (účtovaná zákazníkovi)
SHIPPING_REAL_NET = SHIPPING_REAL_GROSS / (1 + VAT_RATE)  # Reálná doprava bez DPH (~9,2414 CHF)
FREE_SHIPPING_THRESHOLD = 99.0  # Hranice produktové hodnoty pro dopravu zdarma (CHF)


def _calculate_total_price(customs_value: float | None) -> float | None:
    """
    Vypočítá celkovou cenu zásilky v CHF, kterou zákazník skutečně zaplatil.

    Zásilkovna vždy deklaruje v Customs Value: produkty (bez DPH) + 9 CHF doprava.
    Proto nejprve odečteme 9 CHF, abychom získali čistou produktovou hodnotu.
    Threshold 99 CHF se porovnává s touto čistou hodnotou.

    - product_value > 99 CHF: doprava zdarma → product_value × 1,081
    - product_value ≤ 99 CHF: zákazník platí 9,99 CHF dopravu →
          (product_value + 9,99 / 1,081) × 1,081
    """
    if customs_value is None:
        return None
    product_value = customs_value - SHIPPING_IN_CUSTOMS
    if product_value > FREE_SHIPPING_THRESHOLD:
        return round(product_value * (1 + VAT_RATE), 2)
    return round((product_value + SHIPPING_REAL_NET) * (1 + VAT_RATE), 2)

CUSTOMS_VALUE_HEADER = "Customs Value (CHF)"


TRACKING_HEADER_NAMES = {"tracking number", "tracking no", "trasovací číslo"}


def _find_tracking_column(sheet) -> tuple[int, int]:
    """
    Najde sloupec s trasovacími čísly. Nejdřív hledá podle záhlaví,
    pak fallback na regex pattern v datech.
    Vrátí (index_sloupce, index_prvního_datového_řádku).
    """
    # 1) Hledání podle záhlaví
    for row in sheet.iter_rows(max_row=5):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() in TRACKING_HEADER_NAMES:
                return cell.column, cell.row + 1

    # 2) Fallback: hledání podle patternu v datech
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and TRACKING_PATTERN.match(str(cell.value).strip()):
                return cell.column, cell.row
    raise ValueError(
        "Nepodařilo se najít sloupec s trasovacími čísly. "
        "Trasovací čísla musí být ve formátu Z + číslice (např. Z1234567890) "
        "nebo jen číslice (např. 1234567890)."
    )


def _find_customs_value_column(sheet, header_row: int) -> int | None:
    """
    Hledá sloupec 'Customs Value (CHF)' v záhlavním řádku.
    Vrátí index sloupce nebo None pokud není nalezen.
    """
    for cell in sheet[header_row]:
        if cell.value and str(cell.value).strip() == CUSTOMS_VALUE_HEADER:
            return cell.column
    return None

def _insert_order_column(sheet, tracking_col_idx: int) -> tuple[int, int]:
    """Vloží dva nové sloupce za sloupec s trasovacími čísly: číslo objednávky a celková cena."""
    order_col_idx = tracking_col_idx + 1
    sheet.insert_cols(order_col_idx, amount=2)
    price_col_idx = order_col_idx + 1
    return order_col_idx, price_col_idx

def process_excel(file_bytes: bytes, api_password: str) -> tuple[bytes, dict]:
    """
    Zpracuje Excel soubor: přidá sloupec s čísly objednávek.
    Vrátí (zpracované_bytes, statistiky).
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    sheet = wb.active

    tracking_col_idx, first_data_row = _find_tracking_column(sheet)

    # Najdeme sloupec s Customs Value — musí být v záhlaví PŘED vložením nových sloupců
    header_row = first_data_row - 1 if first_data_row > 1 else None
    customs_col_idx = _find_customs_value_column(sheet, header_row) if header_row else None

    order_col_idx, price_col_idx = _insert_order_column(sheet, tracking_col_idx)

    # Po vložení sloupců posuneme index customs value, pokud byl za tracking sloupcem
    if customs_col_idx and customs_col_idx > tracking_col_idx:
        customs_col_idx += 2

    order_col_letter = get_column_letter(order_col_idx)
    price_col_letter = get_column_letter(price_col_idx)

    # Záhlaví sloupců
    if first_data_row > 1:
        for col_idx, header, color in [
            (order_col_idx, ORDER_COLUMN_HEADER, "D9E1F2"),
            (price_col_idx, PRICE_COLUMN_HEADER, "D9F2E1"),
        ]:
            cell = sheet.cell(row=first_data_row - 1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, name="Arial")
            cell.fill = PatternFill("solid", start_color=color)
            cell.alignment = Alignment(horizontal="center")

    sheet.column_dimensions[order_col_letter].width = 22
    sheet.column_dimensions[price_col_letter].width = 18

    # Zahřejeme SOAP klient před paralelním zpracováním (načte WSDL jednou)
    get_client()

    # Sesbíráme řádky s trasovacími čísly
    rows_to_process = []
    for row_idx in range(first_data_row, sheet.max_row + 1):
        val = sheet.cell(row=row_idx, column=tracking_col_idx).value
        if val:
            rows_to_process.append((row_idx, str(val)))

    # Paralelní volání API (jen pro čísla objednávek)
    results: dict[int, dict] = {}
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_row = {
            executor.submit(get_packet_data, api_password, tracking): row_idx
            for row_idx, tracking in rows_to_process
        }
        for future in as_completed(future_to_row):
            row_idx = future_to_row[future]
            results[row_idx] = future.result()

    # Zápis výsledků do Excelu
    stats = {"total": len(rows_to_process), "found": 0, "not_found": 0}
    for row_idx, _ in rows_to_process:
        data = results.get(row_idx, {})
        order_number = data.get("order_number")

        # Hodnota zásilky — čteme přímo z Excelu ze sloupce "Customs Value (CHF)"
        raw_value = sheet.cell(row=row_idx, column=customs_col_idx).value if customs_col_idx else None
        try:
            customs_value = float(raw_value) if raw_value is not None else None
        except (TypeError, ValueError):
            customs_value = None
        total_price = _calculate_total_price(customs_value)

        order_cell = sheet.cell(row=row_idx, column=order_col_idx)
        price_cell = sheet.cell(row=row_idx, column=price_col_idx)

        if order_number:
            order_cell.value = order_number
            order_cell.font = Font(name="Arial")
            stats["found"] += 1
        else:
            order_cell.value = "—"
            order_cell.font = Font(name="Arial", color="999999")
            stats["not_found"] += 1

        if total_price is not None:
            price_cell.value = total_price
            price_cell.font = Font(name="Arial")
            price_cell.number_format = '#,##0.00 "CHF"'
        else:
            price_cell.value = "—"
            price_cell.font = Font(name="Arial", color="999999")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), stats
